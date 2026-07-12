"""Excel (.xlsx) tools: read workbooks into rows, write rows into workbooks.

Both sides ride the workflow file store: ``excel_read`` consumes a file
reference (e.g. from ``file_download``), ``excel_write`` produces one (so
run results show a download link and later nodes can pass it on).
"""

from __future__ import annotations

import io
from datetime import date, datetime, time
from typing import Any

from genxai.core.files import get_file_store, is_file_ref
from genxai.tools.base import Tool, ToolCategory, ToolMetadata, ToolParameter

XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
MAX_READ_ROWS = 10_000


def _json_safe(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return value


class ExcelReadTool(Tool):
    """Read a sheet of an .xlsx workbook into a list of row objects."""

    def __init__(self) -> None:
        super().__init__(
            metadata=ToolMetadata(
                name="excel_read",
                description=(
                    "Read an Excel (.xlsx) file from the workflow file store "
                    "into rows — combine with file_download to ingest "
                    "spreadsheets from URLs"
                ),
                category=ToolCategory.FILE,
                tags=["excel", "xlsx", "spreadsheet", "file", "read"],
                version="1.0.0",
            ),
            parameters=[
                ToolParameter(
                    name="file",
                    type="object",
                    description="File reference (or id) of the .xlsx workbook",
                    required=True,
                ),
                ToolParameter(
                    name="sheet",
                    type="string",
                    description="Sheet name (default: the first sheet)",
                    required=False,
                ),
                ToolParameter(
                    name="header_row",
                    type="boolean",
                    description=(
                        "Treat the first row as column names (default true); "
                        "otherwise columns are col_1, col_2, …"
                    ),
                    required=False,
                    default=True,
                ),
                ToolParameter(
                    name="max_rows",
                    type="number",
                    description="Maximum data rows to return",
                    required=False,
                    default=MAX_READ_ROWS,
                    min_value=1,
                    max_value=MAX_READ_ROWS,
                ),
            ],
        )

    async def _execute(self, **kwargs: Any) -> dict[str, Any]:
        import openpyxl

        ref = kwargs["file"]
        if not (is_file_ref(ref) or isinstance(ref, str)):
            raise ValueError("'file' must be a file reference or file id")
        data = get_file_store().read_bytes(ref)
        workbook = openpyxl.load_workbook(
            io.BytesIO(data), read_only=True, data_only=True
        )
        try:
            sheet_names = workbook.sheetnames
            requested = kwargs.get("sheet")
            if requested and requested not in sheet_names:
                raise ValueError(
                    f"Sheet {requested!r} not found — workbook has {sheet_names}"
                )
            sheet = workbook[requested] if requested else workbook[sheet_names[0]]

            use_header = kwargs.get("header_row", True)
            max_rows = int(kwargs.get("max_rows") or MAX_READ_ROWS)
            rows_iter = sheet.iter_rows(values_only=True)

            columns: list[str] = []
            if use_header:
                header = next(rows_iter, None)
                if header is None:
                    return {"sheet": sheet.title, "sheets": sheet_names, "rows": [],
                            "row_count": 0, "truncated": False}
                columns = [
                    str(cell) if cell is not None else f"col_{index + 1}"
                    for index, cell in enumerate(header)
                ]

            rows: list[dict[str, Any]] = []
            truncated = False
            for values in rows_iter:
                if len(rows) >= max_rows:
                    truncated = True
                    break
                if not columns:
                    columns = [f"col_{index + 1}" for index in range(len(values))]
                if all(value is None for value in values):
                    continue
                rows.append(
                    {
                        columns[index] if index < len(columns) else f"col_{index + 1}":
                            _json_safe(value)
                        for index, value in enumerate(values)
                    }
                )
        finally:
            workbook.close()

        return {
            "sheet": sheet.title,
            "sheets": sheet_names,
            "rows": rows,
            "row_count": len(rows),
            "truncated": truncated,
        }


class ExcelWriteTool(Tool):
    """Write a list of row objects into an .xlsx workbook in the file store."""

    def __init__(self) -> None:
        super().__init__(
            metadata=ToolMetadata(
                name="excel_write",
                description=(
                    "Write rows (a list of objects, e.g. a query result or "
                    "{{ dataset.data.rows }}) into an Excel (.xlsx) file and "
                    "return its file reference"
                ),
                category=ToolCategory.FILE,
                tags=["excel", "xlsx", "spreadsheet", "file", "write"],
                version="1.0.0",
            ),
            parameters=[
                ToolParameter(
                    name="rows",
                    type="array",
                    description="Rows to write — a list of objects",
                    required=True,
                ),
                ToolParameter(
                    name="name",
                    type="string",
                    description="File name (default data.xlsx)",
                    required=False,
                    default="data.xlsx",
                ),
                ToolParameter(
                    name="sheet",
                    type="string",
                    description="Sheet name (default Sheet1)",
                    required=False,
                    default="Sheet1",
                ),
            ],
        )

    async def _execute(self, **kwargs: Any) -> dict[str, Any]:
        import openpyxl

        rows = kwargs["rows"]
        if isinstance(rows, dict):
            rows = [rows]
        if not isinstance(rows, list) or not rows:
            raise ValueError("'rows' must be a non-empty list of objects")

        columns: list[str] = []
        for row in rows:
            if not isinstance(row, dict):
                raise ValueError("every row must be an object")
            for key in row:
                if key not in columns:
                    columns.append(str(key))

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = str(kwargs.get("sheet") or "Sheet1")
        sheet.append(columns)
        for row in rows:
            sheet.append(
                [
                    value if isinstance(value, (int, float, str, bool)) or value is None
                    else str(value)
                    for value in (row.get(column) for column in columns)
                ]
            )

        buffer = io.BytesIO()
        workbook.save(buffer)

        name = str(kwargs.get("name") or "data.xlsx")
        if not name.lower().endswith(".xlsx"):
            name = f"{name}.xlsx"
        ref = get_file_store().save_bytes(
            buffer.getvalue(), name=name, media_type=XLSX_MEDIA_TYPE
        )
        return {"file": ref, "rows": len(rows), "columns": columns}
